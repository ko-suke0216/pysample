# coding: UTF-8
import lxml
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import datetime
import openpyxl

def get_bangumi(file):
    # URL関連
    url = "https://www.home-tv.co.jp/programtable/"
 
    # ヘッドレスモードを設定
    options = Options()
    options.add_argument("--headless")
 
    # WebDriverを設定
    CHROME_PATH = "C:\\work\\99_tool\\chromedriver_win32\\chromedriver.exe"
    driver = webdriver.Chrome(executable_path=CHROME_PATH, options=options)
 
    # htmlを取得
    driver.get(url)
 
    # soupオブジェクトを作成
    soup = BeautifulSoup(driver.page_source, "lxml")
 
    # 番組情報を取得して日付、時刻、番組タイトル、番組概要を出力
    # 出力ファイル設定
    now = datetime.datetime.now()
    output = "C:\\work\\02_Uipath\\" + "bangumi_{0:%Y%m%d%H%M%S}.xlsx".format(now)
    # ワークブック新規作成
    wb = openpyxl.Workbook()
    ws = wb.active

    # 行数
    row_num = 1
    for b_box in soup.find_all("div", class_="broadcaster_box"):
        for d_prog in b_box.find_all("li", attrs={"data-program": "番組枠"}):
            tempdate = b_box.get("data-yyyymmdd")
            strdate = tempdate[0:8]
            strtime = d_prog.find("span", class_="time")
            strtitle = d_prog.find("span", class_="title")
            strdt = d_prog.find("span", class_="dt ellipsis")
            # 時刻が存在しない場合は次へ
            if strtitle is None:
                continue
            if strdate is not None:
                # 日付
                ws.cell(row=row_num, column=1, value=strdate[0:8])
            if strtime is not None:
                # 時刻
                ws.cell(row=row_num, column=2, value=strtime.text)
            if strtitle is not None:
                # 番組タイトル
                ws.cell(row=row_num, column=3, value=strtitle.text)
            if strdt is not None:
                # 番組概要
                ws.cell(row=row_num, column=4, value=strdt.text)
            # 行数をインクリメント
            row_num += 1

    # ワークブックを保存
    wb.save(output)
    # Chromeドライバーを終了
    driver.close()

if __name__ == '__main__':
    #   # arguments
    #   argvs = sys.argv
    #   ## check
    #   if len(argvs) != 3:
    #       print("Usage: python scraping.py [url] [output]")
    #       exit()
    #   url = argvs[1]
    #   output_name = argvs[2]
    url = "https://www.home-tv.co.jp/programtable/"
    file = "bangumi.txt"

    get_bangumi(file)