# endording UTF-8
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
import time
from time import sleep
import datetime
import openpyxl
import os
import logging.config


def get_timesale():
    # ログ設定
    ini_name = os.path.join(os.path.dirname(__file__), "config\\logging.ini")
    logging.config.fileConfig(ini_name)
    logger = logging.getLogger("testLog")
    logger.info("処理を開始します")

    # URL
    URL = "https://www.amazon.co.jp/gp/goldbox?ref_=nav_cs_gb"

    # ヘッドレスモードを設定
    options = Options()
    options.add_argument("--headless")

    # webdriverを設定
    CHROME_PATH = os.path.join(os.path.dirname(__file__), "driver\\chromedriver.exe")
    driver = webdriver.Chrome(executable_path=CHROME_PATH, options=options)

    # htmlを取得
    driver.get(URL)

    # 出力ファイル設定
    now = datetime.datetime.now()
    output = "C:\\work\\99_tool\\" + "amazon_{0:%Y%m%d%H%M%S}.xlsx".format(now)
    # ワークブック新規作成
    wb = openpyxl.Workbook()
    ws = wb.active
    row_num = 1

    # 要素があれば継続
    while True:
        try:
            # soupオブジェクトを作成
            soup = BeautifulSoup(driver.page_source, "lxml")
            # タイムセール商品を取得
            for hinmoku in soup.find_all("div", class_="a-section layer"):
                price = hinmoku.find("span", class_="a-size-medium inlineBlock unitLineHeight dealPriceText")
                product_name = hinmoku.find("a", class_="a-size-base a-link-normal dealTitleTwoLine singleCellTitle autoHeight")
                if hinmoku is None:
                    continue
                if price is None:
                    continue
                if product_name is None:
                    continue
                logger.debug(price.text.strip())
                ws.cell(row=row_num, column=1, value=price.text.strip())
                logger.debug(product_name.text.strip())
                logger.debug(product_name.get("href").strip())
                ws.cell(row=row_num, column=2, value=product_name.text.strip())
                ws.cell(row=row_num, column=3, value=product_name.get("href").strip())
                row_num += 1
            nextpage = driver.find_element_by_partial_link_text("次へ")
            # 次へボタンクリック
            nextpage.click()
            # 読み込みのため5秒間待機
            sleep(5)
            # とりあえず1000件で終了
            if row_num >= 100:
                logger.info("件数が上限に達したため、このへんでやめます")
                break
        except Exception as e:
            logger.info("例外が発生しました")
            logger.info(e)
            # 次へボタンがなければループから抜ける
            break
    
    # ワークブックを保存
    wb.save(output)
    # 終了
    driver.quit()
    # 終了ログ
    logger.info("処理を終了します")

if __name__ == "__main__":

    get_timesale()